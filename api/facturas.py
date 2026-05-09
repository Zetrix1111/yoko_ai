"""
api/facturas.py — dispatcher de Facturas Inteligentes.

Único archivo serverless del módulo. Consolida múltiples acciones en un
solo dispatcher por `?action=` y método HTTP libre (cada action sabe
qué método espera y parsea el body acorde).

Acciones del flujo web (UI clásica):
  POST   /api/facturas?action=procesar      → multipart, procesa N archivos
  PUT    /api/facturas?action=actualizar    → JSON, auto-save de ediciones
  GET    /api/facturas?action=recuperar     → ?proceso_id=…, recupera proceso
  DELETE /api/facturas?action=eliminar-fila → JSON, borra una factura
  POST   /api/facturas?action=concar        → genera y descarga el Excel CONCAR

Acciones consumidas por el agent (Anthropic Managed Agents → custom tools):
  POST   /api/facturas?action=procesar-chat   → JSON+base64, equivalente a procesar
  POST   /api/facturas?action=recuperar-chat  → JSON {proceso_id}, equivalente a recuperar
  POST   /api/facturas?action=download-chat   → JSON {proceso_id}, alias de concar

Todas validan JWT en el dispatcher; `empresa_id` se extrae del token.
"""

import base64
import cgi
import io
import json
import os
import sys
import time
import uuid
from datetime import datetime
from http.server import BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs


_HERE = os.path.dirname(os.path.abspath(__file__))
if _HERE not in sys.path:
    sys.path.insert(0, _HERE)

from _lib import auth, config_loader                                 # noqa: E402
from _lib.auth import AuthError                                      # noqa: E402
from _lib.airtable_client import AirtableError                       # noqa: E402
from _lib.facturas_processor import process_multiple_files           # noqa: E402
from _lib.db_manager import (                                        # noqa: E402
    init_db,
    save_proceso,
    get_proceso,
    update_factura,
    delete_factura,
)
from _lib.contabilidad import (                                      # noqa: E402
    get_contabilidad_config,
    factura_a_filas_excel,
    CONCAR_EXCEL_HEADERS,
)


# ─────────────────────────────────────────────────────────────────────────
# Acciones — invocadas tras validar JWT en el dispatcher
# ─────────────────────────────────────────────────────────────────────────

def _procesar(req, empresa_id: str) -> None:
    """
    POST multipart/form-data — procesa N archivos en paralelo.

    Campos del body:
      - tipo:       "compra" | "venta"
      - mes:        "YYYY-MM"
      - mes_label:  "Mayo 2026"
      - dni:        DNI del usuario
      - files:      uno o más archivos (campo repetible)

    Response: {ok, proceso_id, empresa_id, facturas, errores, timestamp}
    """
    try:
        init_db()

        content_type   = req.headers.get("Content-Type", "")
        content_length = int(req.headers.get("Content-Length", 0))
        raw_body       = req.rfile.read(content_length)

        environ = {
            "REQUEST_METHOD": "POST",
            "CONTENT_TYPE":   content_type,
            "CONTENT_LENGTH": str(content_length),
        }
        fs = cgi.FieldStorage(
            fp=io.BytesIO(raw_body),
            environ=environ,
            keep_blank_values=True,
        )

        # Campos de texto del body multipart.
        def _str(name: str, default: str = "") -> str:
            if name not in fs:
                return default
            v = fs.getvalue(name)
            return v.decode("utf-8") if isinstance(v, bytes) else (v or default)

        tipo = _str("tipo", "compra")
        mes  = _str("mes", "")

        # Archivos: el campo "files" puede aparecer una o N veces.
        files = []
        if "files" in fs:
            file_items = fs["files"]
            if not isinstance(file_items, list):
                file_items = [file_items]
            for item in file_items:
                if hasattr(item, "filename") and hasattr(item, "file"):
                    filename = item.filename or "unknown"
                    file_bytes = item.file.read()
                    files.append((filename, file_bytes))

        if not files:
            return req._json(400, {"error": "No se recibieron archivos."})
        if len(files) > 50:
            return req._json(400, {"error": "Máximo 50 archivos por lote."})

        api_key = os.environ.get("OPENAI_API_KEY")
        if not api_key:
            return req._json(500, {"error": "OPENAI_API_KEY no configurada."})

        proceso_id = f"proc-{uuid.uuid4().hex[:12]}"

        resultado = process_multiple_files(
            files=files,
            tipo=tipo,
            mes=mes,
            api_key=api_key,
        )

        save_proceso(proceso_id, empresa_id, resultado["facturas"])

        return req._json(200, {
            "ok":         True,
            "proceso_id": proceso_id,
            "empresa_id": empresa_id,
            "facturas":   resultado["facturas"],
            "errores":    resultado["errores"],
            "timestamp":  time.time(),
        })

    except ValueError as e:
        print(f"[facturas/procesar] ValueError: {e}", file=sys.stderr)
        return req._json(400, {"error": str(e)})
    except Exception as e:
        print(f"[facturas/procesar] Error: {type(e).__name__}: {e}", file=sys.stderr)
        return req._json(500, {"error": "Error al procesar facturas."})


def _actualizar(req, empresa_id: str) -> None:
    """
    PUT JSON — auto-save de ediciones del usuario.

    Body: {proceso_id, facturas: [...]}
    Response: {ok, updated_count}
    """
    try:
        length = int(req.headers.get("Content-Length", 0))
        if length == 0:
            return req._json(400, {"error": "Body vacío."})

        try:
            body = json.loads(req.rfile.read(length))
        except json.JSONDecodeError:
            return req._json(400, {"error": "JSON inválido."})

        proceso_id = body.get("proceso_id")
        facturas   = body.get("facturas", [])

        if not proceso_id:
            return req._json(400, {"error": "proceso_id requerido."})

        # Cross-tenant guard: el proceso debe existir y ser del tenant.
        if not get_proceso(proceso_id, empresa_id):
            return req._json(404, {"error": "Proceso no encontrado."})

        updated_count = sum(
            1 for f in facturas
            if update_factura(proceso_id, empresa_id, f)
        )

        return req._json(200, {"ok": True, "updated_count": updated_count})

    except Exception as e:
        print(f"[facturas/actualizar] Error: {type(e).__name__}: {e}", file=sys.stderr)
        return req._json(500, {"error": "Error al actualizar."})


def _recuperar(req, empresa_id: str) -> None:
    """
    GET ?proceso_id=… — recupera un proceso guardado previamente.

    Response: {ok, proceso_id, facturas, timestamp} o 404 si no existe.
    """
    try:
        proceso_id = (parse_qs(urlparse(req.path).query).get("proceso_id") or [""])[0]
        if not proceso_id:
            return req._json(400, {"error": "proceso_id requerido en query."})

        proceso = get_proceso(proceso_id, empresa_id)
        if not proceso:
            return req._json(404, {"error": "Proceso no encontrado o expirado."})

        return req._json(200, {
            "ok":         True,
            "proceso_id": proceso_id,
            "facturas":   proceso["facturas"],
            "timestamp":  proceso["timestamp"],
        })

    except Exception as e:
        print(f"[facturas/recuperar] Error: {type(e).__name__}: {e}", file=sys.stderr)
        return req._json(500, {"error": "Error al recuperar proceso."})


def _eliminar_fila(req, empresa_id: str) -> None:
    """
    DELETE JSON — elimina una factura del proceso.

    Body: {proceso_id, factura_id}
    Response: {ok}
    """
    try:
        length = int(req.headers.get("Content-Length", 0))
        if length == 0:
            return req._json(400, {"error": "Body vacío."})

        try:
            body = json.loads(req.rfile.read(length))
        except json.JSONDecodeError:
            return req._json(400, {"error": "JSON inválido."})

        proceso_id = body.get("proceso_id")
        factura_id = body.get("factura_id")
        if not proceso_id or not factura_id:
            return req._json(400, {"error": "proceso_id y factura_id requeridos."})

        # Cross-tenant guard: el proceso debe existir y ser del tenant.
        if not get_proceso(proceso_id, empresa_id):
            return req._json(404, {"error": "Proceso no encontrado."})

        if not delete_factura(proceso_id, empresa_id, factura_id):
            return req._json(404, {"error": "Factura no encontrada."})

        return req._json(200, {"ok": True})

    except Exception as e:
        print(f"[facturas/eliminar-fila] Error: {type(e).__name__}: {e}", file=sys.stderr)
        return req._json(500, {"error": "Error al eliminar."})


def _concar(req, empresa_id: str) -> None:
    """
    POST JSON — genera el archivo Excel CONCAR para descarga.

    Body: {proceso_id}

    Lee las facturas validadas de SQLite, resuelve la config contable de
    la empresa, genera 2-3 filas por factura (DEBE gasto / DEBE IGV / HABER
    cxp) y construye un .xlsx con openpyxl: filas 1-3 con headers azules
    (replicando la plantilla CONCAR del usuario), datos desde fila 4.

    Response: binary stream `application/vnd.openxmlformats-officedocument.spreadsheetml.sheet`
    con `Content-Disposition: attachment; filename="CONCAR_<proceso>.xlsx"`.
    El frontend lo guarda como blob y dispara la descarga.
    """
    try:
        length = int(req.headers.get("Content-Length", 0))
        if length == 0:
            return req._json(400, {"error": "Body vacío."})

        try:
            body = json.loads(req.rfile.read(length))
        except json.JSONDecodeError:
            return req._json(400, {"error": "JSON inválido."})

        proceso_id = body.get("proceso_id")
        if not proceso_id:
            return req._json(400, {"error": "proceso_id requerido."})

        # 1) Cargar proceso desde SQLite (cross-tenant guard via empresa_id).
        proceso = get_proceso(proceso_id, empresa_id)
        if not proceso:
            return req._json(404, {"error": "Proceso no encontrado o expirado."})
        facturas = proceso.get("facturas") or []
        if not facturas:
            return req._json(400, {"error": "El proceso no tiene facturas."})

        # 2) Cargar config contable de la empresa (sistema_contable + overrides).
        try:
            full_config = config_loader.load_full_config(empresa_id)
        except AirtableError as e:
            print(f"[facturas/concar] AirtableError config: {e}", file=sys.stderr)
            full_config = {"empresa": {}}
        empresa_data = full_config.get("empresa") or {}
        contab = get_contabilidad_config(empresa_data)

        # 3) Generar todas las filas (cada factura → 2-3 filas).
        fecha_hoy = datetime.now().strftime("%d/%m/%Y")
        todas_las_filas: list = []
        for factura in facturas:
            todas_las_filas.extend(factura_a_filas_excel(factura, contab, fecha_hoy))

        print(
            f"[facturas/concar] {proceso_id}: {len(facturas)} facturas → "
            f"{len(todas_las_filas)} filas en el Excel",
            file=sys.stderr,
        )

        # 4) Generar .xlsx con openpyxl. Headers en filas 1-3, datos desde fila 4.
        try:
            xlsx_bytes = _generar_xlsx_concar(todas_las_filas)
        except Exception as e:
            print(f"[facturas/concar] Error generando xlsx: {e}", file=sys.stderr)
            return req._json(500, {"error": "Error al generar el Excel."})

        # 5) Stream binary download.
        filename = f"CONCAR_{proceso_id}.xlsx"
        req.send_response(200)
        req.send_header(
            "Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        req.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        req.send_header("Content-Length", str(len(xlsx_bytes)))
        req.end_headers()
        req.wfile.write(xlsx_bytes)

    except Exception as e:
        print(f"[facturas/concar] Error: {type(e).__name__}: {e}", file=sys.stderr)
        return req._json(500, {"error": "Error al generar el archivo CONCAR."})


def _generar_xlsx_concar(filas: list) -> bytes:
    """
    Construye un .xlsx en memoria con:
      - Fila 1: headers azules (cabecera de columna).
      - Fila 2: descripciones de validación.
      - Fila 3: tamaño/formato de cada columna.
      - Fila 4 en adelante: las filas de datos generadas por
        factura_a_filas_excel (keys = letras de columna).

    Devuelve el contenido binario del .xlsx listo para servir.
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CONCAR"

    # Fila 1: nombres de columna con cabecera azul + texto blanco bold.
    fill_blue   = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    font_white  = Font(color="FFFFFF", bold=True)
    font_small  = Font(size=9, italic=True)
    align_wrap  = Alignment(wrap_text=True, vertical="top")

    for col, header in CONCAR_EXCEL_HEADERS["row1"].items():
        cell = ws[f"{col}1"]
        cell.value = header
        cell.fill = fill_blue
        cell.font = font_white
        cell.alignment = align_wrap

    # Fila 2: descripciones de validación (amarillo claro).
    for col, desc in CONCAR_EXCEL_HEADERS["row2"].items():
        cell = ws[f"{col}2"]
        cell.value = desc
        cell.fill = fill_yellow
        cell.font = font_small
        cell.alignment = align_wrap

    # Fila 3: tamaño/formato (texto pequeño en gris).
    for col, fmt in CONCAR_EXCEL_HEADERS["row3"].items():
        cell = ws[f"{col}3"]
        cell.value = fmt
        cell.font = font_small
        cell.alignment = align_wrap

    # Filas 4+ : datos.
    for row_idx, fila in enumerate(filas, start=4):
        for col_letter, value in fila.items():
            ws[f"{col_letter}{row_idx}"] = value

    # Anchos razonables para las columnas más comunes.
    anchos = {
        "A":  6,  "B": 10, "C": 14, "D": 14, "E": 12,
        "F": 40,  "G": 14, "H": 12, "I": 16, "J": 14,
        "K": 14, "L": 18, "M": 16, "N": 12, "O": 14,
        "P": 14, "Q": 14, "R": 12, "S": 18, "T": 14,
        "U": 14, "W": 30, "AO": 12,
    }
    for col, width in anchos.items():
        ws.column_dimensions[col].width = width

    # Altura mayor para fila 1 y 2 (headers con texto largo).
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 60

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _procesar_chat(req, empresa_id: str) -> None:
    """
    POST JSON — variante "chat-friendly" de procesar.

    En vez de multipart, recibe los archivos en base64 dentro del JSON, que
    es como Anthropic Managed Agents le pasa el lote al custom tool
    `yoko_procesar_archivos`. Reusa exactamente la misma lógica de extracción
    que `_procesar`.

    Body:
      {
        "tipo":  "compra" | "venta",
        "mes":   "YYYY-MM",
        "files": [{"filename": "...", "content_b64": "..."}]   # ≤ 50
      }

    Response: {ok, proceso_id, empresa_id, facturas, errores, alertas, timestamp}
    """
    try:
        init_db()

        length = int(req.headers.get("Content-Length", 0))
        if length == 0:
            return req._json(400, {"error": "Body vacío."})
        try:
            body = json.loads(req.rfile.read(length))
        except json.JSONDecodeError:
            return req._json(400, {"error": "JSON inválido."})

        tipo     = (body.get("tipo") or "compra").strip().lower()
        mes      = (body.get("mes") or "").strip()
        files_in = body.get("files") or []

        if tipo not in ("compra", "venta"):
            return req._json(400, {"error": "tipo debe ser 'compra' o 'venta'."})
        if not isinstance(files_in, list) or not files_in:
            return req._json(400, {"error": "files debe ser una lista no vacía."})
        if len(files_in) > 50:
            return req._json(400, {"error": "Máximo 50 archivos por lote."})

        # Decodificar base64 → list[(filename, bytes)]
        files: list[tuple[str, bytes]] = []
        for i, item in enumerate(files_in):
            if not isinstance(item, dict):
                return req._json(400, {
                    "error": f"files[{i}] debe ser objeto con filename+content_b64.",
                })
            fname = (item.get("filename") or "").strip()
            content_b64 = item.get("content_b64") or ""
            if not fname or not content_b64:
                return req._json(400, {
                    "error": f"files[{i}] requiere filename y content_b64 no vacíos.",
                })
            try:
                fbytes = base64.b64decode(content_b64, validate=True)
            except (ValueError, base64.binascii.Error):
                return req._json(400, {"error": f"files[{i}] content_b64 no es base64 válido."})
            if not fbytes:
                return req._json(400, {"error": f"files[{i}] vacío tras decodificar."})
            files.append((fname, fbytes))

        api_key = os.environ.get("OPENAI_API_KEY")
        if not api_key:
            return req._json(500, {"error": "OPENAI_API_KEY no configurada."})

        proceso_id = f"proc-{uuid.uuid4().hex[:12]}"

        resultado = process_multiple_files(
            files=files,
            tipo=tipo,
            mes=mes,
            api_key=api_key,
        )
        save_proceso(proceso_id, empresa_id, resultado["facturas"])

        return req._json(200, {
            "ok":         True,
            "proceso_id": proceso_id,
            "empresa_id": empresa_id,
            "facturas":   resultado["facturas"],
            "errores":    resultado["errores"],
            "alertas":    [],
            "timestamp":  time.time(),
        })

    except ValueError as e:
        print(f"[facturas/procesar-chat] ValueError: {e}", file=sys.stderr)
        return req._json(400, {"error": str(e)})
    except Exception as e:
        print(f"[facturas/procesar-chat] Error: {type(e).__name__}: {e}", file=sys.stderr)
        return req._json(500, {"error": "Error al procesar facturas."})


def _recuperar_chat(req, empresa_id: str) -> None:
    """
    POST JSON — variante de recuperar pensada para el agent.

    Diferencia con `_recuperar` (que es GET con ?proceso_id=…): el agent
    llama esto en POST con body JSON, que es la forma natural en la que
    Anthropic invoca un custom tool.

    Body: {"proceso_id": "..."}
    Response: {ok, proceso_id, facturas, timestamp}
    """
    try:
        length = int(req.headers.get("Content-Length", 0))
        if length == 0:
            return req._json(400, {"error": "Body vacío."})
        try:
            body = json.loads(req.rfile.read(length))
        except json.JSONDecodeError:
            return req._json(400, {"error": "JSON inválido."})

        proceso_id = body.get("proceso_id")
        if not proceso_id:
            return req._json(400, {"error": "proceso_id requerido."})

        proceso = get_proceso(proceso_id, empresa_id)
        if not proceso:
            return req._json(404, {"error": "Proceso no encontrado o expirado."})

        return req._json(200, {
            "ok":         True,
            "proceso_id": proceso_id,
            "facturas":   proceso["facturas"],
            "timestamp":  proceso["timestamp"],
        })

    except Exception as e:
        print(f"[facturas/recuperar-chat] Error: {type(e).__name__}: {e}", file=sys.stderr)
        return req._json(500, {"error": "Error al recuperar proceso."})


_ACTIONS = {
    # Flujo web (UI clásica)
    "procesar":      _procesar,
    "actualizar":    _actualizar,
    "recuperar":     _recuperar,
    "eliminar-fila": _eliminar_fila,
    "concar":        _concar,
    # Flujo chat (Anthropic Managed Agents → custom tools)
    "procesar-chat":  _procesar_chat,
    "recuperar-chat": _recuperar_chat,
    "download-chat":  _concar,   # alias: misma lógica que ?action=concar
}


# ─────────────────────────────────────────────────────────────────────────
# Dispatcher — método-agnóstico, rutea solo por ?action=
# ─────────────────────────────────────────────────────────────────────────

class handler(BaseHTTPRequestHandler):

    def do_POST(self) -> None:    return self._dispatch()
    def do_PUT(self) -> None:     return self._dispatch()
    def do_GET(self) -> None:     return self._dispatch()
    def do_DELETE(self) -> None:  return self._dispatch()

    def _dispatch(self) -> None:
        action = (parse_qs(urlparse(self.path).query).get("action") or [""])[0]
        fn = _ACTIONS.get(action)
        if fn is None:
            return self._json(400, {
                "error": f"action inválida. Use: {sorted(_ACTIONS)}",
            })

        try:
            try:
                auth_payload = auth.require_auth(self.headers)
            except AuthError as e:
                return self._json(e.status, {"error": str(e)})
            empresa_id = auth_payload["empresa_id"]
            return fn(self, empresa_id)
        except Exception as e:
            print(f"[facturas/{action}] Error: {type(e).__name__}: {e}", file=sys.stderr)
            return self._json(500, {"error": "Error interno del servidor."})

    def _json(self, status: int, data: dict) -> None:
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, *args):
        pass
