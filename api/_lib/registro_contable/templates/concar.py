"""
api/_lib/registro_contable/templates/concar.py

Template de salida para el sistema contable CONCAR (Perú).

Este módulo es una copia literal de las constantes y funciones que vivían
en `_lib/contabilidad.py` (CONCAR_DEFAULTS, factura_a_filas_excel,
CONCAR_EXCEL_HEADERS) y `facturas.py` (_generar_xlsx_concar). NO modificar
la lógica — está calibrada en producción contra plantillas reales del
usuario.
"""

import io
from typing import Dict, List


NAME = "concar"
DESCRIPTION = "Registro de compras/ventas formato CONCAR (Perú)"


# ─────────────────────────────────────────────────────────────────────────
# Plan de cuentas y mapeos por defecto
# ─────────────────────────────────────────────────────────────────────────

# Plan de cuentas estándar peruano para CONCAR. Replica las constantes del
# escenario Make.com original. La empresa puede overridear cualquiera de
# estos valores en Config_Empresa.data.contabilidad.
DEFAULTS: Dict = {
    "cuentas": {
        "gasto":      "63/65",   # 63 = servicios prestados por terceros, 65 = otros gastos
        "cxp":        "421201",  # cuentas por pagar comerciales — facturas
        "igv":        "401111",  # IGV crédito fiscal
    },
    # Sub-diario CONCAR por tipo de comprobante (codigo de 2 letras del OCR).
    # Confirmado con el usuario: RH=15, BV=13, todos los demás = 11.
    "sub_diarios": {
        "FT": "11",  # Factura
        "BV": "13",  # Boleta de venta
        "NC": "11",  # Nota de crédito
        "ND": "11",  # Nota de débito
        "RH": "15",  # Recibo por honorarios
        "BA": "11",  # Boleto aéreo
        "TK": "11",  # Ticket
    },
    # Tipo de documento (col R del Excel CONCAR). Usamos el código del OCR
    # directamente — CONCAR los acepta como FT, BV, NC, etc.
    "tipos_doc_codigo": {
        "FT": "FT",
        "BV": "BV",
        "NC": "NC",
        "ND": "ND",
        "RH": "RH",
        "BA": "BA",
        "TK": "TK",
    },
    # Mapeo de moneda OCR → código CONCAR (col E del Excel).
    "monedas_codigo": {
        "PEN": "MN",  # Moneda Nacional (soles)
        "USD": "ME",  # Moneda Extranjera (dólares)
        "EUR": "ER",  # Euros
    },
    "tasa_igv": 18.0,
    "moneda_default": "PEN",
}


# Headers que se escriben en las filas 1-3 del Excel CONCAR.
# Extraídos del template real N_Excel_CONCAR.xlsx.
EXCEL_HEADERS: Dict = {
    # Fila 1: nombres de columna
    "row1": {
        "A": "WE",
        "B": "Sub Diario",
        "C": "Número de Comprobante",
        "D": "Fecha de Comprobante",
        "E": "Código de Moneda",
        "F": "Glosa Principal",
        "G": "Tipo de Cambio",
        "H": "Tipo de Conversión",
        "I": "Flag de Conversión de Moneda",
        "J": "Fecha Tipo de Cambio",
        "K": "Cuenta Contable",
        "L": "Código de Anexo",
        "M": "Código de Centro de Costo",
        "N": "Debe / Haber",
        "O": "Importe Original",
        "P": "Importe en Dólares",
        "Q": "Importe en Soles",
        "R": "Tipo de Documento",
        "S": "Número de Documento",
        "T": "Fecha de Documento",
        "U": "Fecha de Vencimiento",
        "V": "Código de Area",
        "W": "Glosa Detalle",
        "X": "Código de Anexo Auxiliar",
        "Y": "Medio de Pago",
        "Z": "Tipo de Documento de Referencia",
        "AA": "Número de Documento Referencia",
        "AB": "Fecha Documento Referencia",
        "AC": "Nro Máq. Registradora Tipo Doc. Ref.",
        "AD": "Base Imponible Documento Referencia",
        "AE": "IGV Documento Provisión",
        "AF": "Tipo Referencia en estado MQ",
        "AG": "Número Serie Caja Registradora",
        "AH": "Fecha de Operación",
        "AI": "Tipo de Tasa",
        "AJ": "Tasa Detracción/Percepción",
        "AK": "Importe Base Detracción/Percepción Dólares",
        "AL": "Importe Base Detracción/Percepción Soles",
        "AM": "Tipo Cambio para 'F'",
        "AN": "Importe de IGV sin derecho crédito fiscal",
        "AO": "Tasa IGV",
    },
    # Fila 2: descripciones de validación
    "row2": {
        "A": "Contabilidad",
        "B": "Ver T.G. 02",
        "C": "Los dos primeros dígitos son el mes y los otros 4 siguientes un correlativo",
        "E": "Ver T.G. 03",
        "G": "Llenar  solo si Tipo de Conversión es 'C'. Debe estar entre >=0 y <=9999.999999",
        "H": "Solo: 'C'= Especial, 'M'=Compra, 'V'=Venta , 'F' De acuerdo a fecha",
        "I": "Solo: 'S' = Si se convierte, 'N'= No se convierte",
        "J": "Si  Tipo de Conversión 'F'",
        "K": "Debe existir en el Plan de Cuentas",
        "L": "Si Cuenta Contable tiene seleccionado Tipo de Anexo, debe existir en la tabla de Anexos",
        "M": "Si Cuenta Contable tiene habilitado C. Costo, Ver T.G. 05",
        "N": "'D' ó 'H'",
        "O": "Importe original de la cuenta contable. Obligatorio, debe estar entre >=0 y <=99999999999.99",
        "P": "Importe de la Cuenta Contable en Dólares. Obligatorio si Flag de Conversión de Moneda esta en 'N', debe estar entre >=0 y <=99999999999.99",
        "Q": "Importe de la Cuenta Contable en Soles. Obligatorio si Flag de Conversión de Moneda esta en 'N', debe estra entre >=0 y <=99999999999.99",
        "R": "Si Cuenta Contable tiene habilitado el Documento Referencia Ver T.G. 06",
        "S": "Si Cuenta Contable tiene habilitado el Documento Referencia Incluye Serie y Número",
        "T": "Si Cuenta Contable tiene habilitado el Documento Referencia",
        "U": "Si Cuenta Contable tiene habilitada la Fecha de Vencimiento",
        "V": "Si Cuenta Contable tiene habilitada el Area. Ver T.G. 26",
        "X": "Si Cuenta Contable tiene seleccionado Tipo de Anexo Referencia",
        "Y": "Si Cuenta Contable tiene habilitado Tipo Medio Pago. Ver T.G. 'S1'",
        "Z": "Si Tipo de Documento es 'NA' ó 'ND' Ver T.G. 06",
        "AA": "Si Tipo de Documento es 'NC', 'NA' ó 'ND', incluye Serie y Número",
        "AB": "Si Tipo de Documento es 'NC', 'NA' ó 'ND'",
        "AC": "Si Tipo de Documento es 'NC', 'NA' ó 'ND'. Solo cuando el Tipo Documento de Referencia 'TK'",
        "AD": "Si Tipo de Documento es 'NC', 'NA' ó 'ND'",
        "AE": "Si Tipo de Documento es 'NC', 'NA' ó 'ND'",
        "AF": "Si la Cuenta Contable tiene Habilitado Documento Referencia 2 y  Tipo de Documento es 'TK'",
        "AG": "Si la Cuenta Contable teien Habilitado Documento Referencia 2 y  Tipo de Documento es 'TK'",
        "AH": "Si la Cuenta Contable tiene Habilitado Documento Referencia 2. Cuando Tipo de Documento es 'TK', consignar la fecha de emision del ticket",
        "AI": "Si la Cuenta Contable tiene configurada la Tasa:  Si es '1' ver T.G. 28 y '2' ver T.G. 29",
        "AJ": "Si la Cuenta Contable tiene conf. en Tasa:  Si es '1' ver T.G. 28 y '2' ver T.G. 29. Debe estar entre >=0 y <=999.99",
        "AK": "Si la Cuenta Contable tiene configurada la Tasa. Debe ser el importe total del documento y estar entre >=0 y <=99999999999.99",
        "AL": "Si la Cuenta Contable tiene configurada la Tasa. Debe ser el importe total del documento y estar entre >=0 y <=99999999999.99",
        "AM": "Especificar solo si Tipo Conversión es 'F'. Se permite 'M' Compra y 'V' Venta.",
        "AN": "Especificar solo para comprobantes de compras con IGV sin derecho de crédito Fiscal. Se detalle solo en la cuenta 42xxxx",
        "AO": "Obligatorio para comprobantes de compras, valores validos 0,10,18.",
    },
    # Fila 3: Tamaño/Formato de cada columna
    "row3": {
        "A": "Tamaño/Formato",
        "B": "4 Caracteres",
        "C": "6 Caracteres",
        "D": "dd/mm/aaaa",
        "E": "2 Caracteres",
        "F": "40 Caracteres",
        "G": "Numérico 11, 6",
        "H": "1 Caracteres",
        "I": "1 Caracteres",
        "J": "dd/mm/aaaa",
        "K": "12 Caracteres",
        "L": "18 Caracteres",
        "M": "6 Caracteres",
        "N": "1 Carácter",
        "O": "Numérico 14,2",
        "P": "Numérico 14,2",
        "Q": "Numérico 14,2",
        "R": "2 Caracteres",
        "S": "20 Caracteres",
        "T": "dd/mm/aaaa",
        "U": "dd/mm/aaaa",
        "V": "3 Caracteres",
        "W": "30 Caracteres",
        "X": "18 Caracteres",
        "Y": "8 Caracteres",
        "Z": "2 Caracteres",
        "AA": "20 Caracteres",
        "AB": "dd/mm/aaaa",
        "AC": "20 Caracteres",
        "AD": "Numérico 14,2",
        "AE": "Numérico 14,2",
        "AF": "'MQ'",
        "AG": "15 caracteres",
        "AH": "dd/mm/aaaa",
        "AI": "5 Caracteres",
        "AJ": "Numérico 14,2",
        "AK": "Numérico 14,2",
        "AL": "Numérico 14,2",
        "AM": "1 Caracter",
        "AN": "Numérico 14,2",
        "AO": "Numérico 14,2",
    },
}


# Constantes para el formato del Excel.
_DEFAULT_TIPO_CONVERSION = "V"
_DEFAULT_FLAG_CONVERSION = "S"


# ─────────────────────────────────────────────────────────────────────────
# Generador de filas (factura → 2-3 filas A-AO)
# ─────────────────────────────────────────────────────────────────────────

def factura_a_filas(
    factura: Dict, contab: Dict, fecha_hoy: str,
    numero_comprobante: str = "",
) -> List[Dict]:
    """
    Convierte una factura en 2-3 dicts cuyas KEYS son letras de columna
    (A, B, ..., AO) listas para escribir directo al Excel CONCAR.

    Cada dict representa una línea contable:
      - Línea 1: GASTO (DEBE)        — cuenta gasto, base sin IGV, cc = obra_area.
      - Línea 2: IGV CRÉDITO (DEBE)  — cuenta IGV, monto = monto_tributo. Solo si > 0.
      - Línea 3: CXP PROVEEDOR (HABER) — cuenta cxp, total con IGV, anexo = ruc.

    Args:
        factura:           dict con campos OCR-extraídos + obra_area.
        contab:            dict resultado de engine.merge_config(DEFAULTS, overrides).
        fecha_hoy:         string DD/MM/YYYY para columna J ("Fecha Tipo de Cambio").
        numero_comprobante: string para columna C ("Número de Comprobante").
                            Formato esperado MMNNNN (6 chars). El cálculo del
                            valor (mes + correlativo zero-padded por sub_diario)
                            vive en `engine.generate`; acá solo lo usamos en la
                            salida. Default "" deja la columna vacía (back-compat
                            con consumers que llaman sin este arg).
    """
    tipo_doc    = (factura.get("tipo_doc_codigo") or "FT").upper()
    sub_diario  = contab["sub_diarios"].get(tipo_doc, "11")
    tipo_concar = contab["tipos_doc_codigo"].get(tipo_doc, "FT")

    moneda      = (factura.get("moneda") or "PEN").upper()
    moneda_code = contab["monedas_codigo"].get(moneda, "MN")

    monto_total = float(factura.get("monto_total") or 0)
    monto_igv   = float(factura.get("monto_tributo") or 0)
    base        = round(monto_total - monto_igv, 2)

    ruc = (factura.get("ruc") or "").strip()
    cc  = (factura.get("obra_area") or "").strip()

    fecha_emi   = factura.get("fecha_emision") or ""
    vencimiento = factura.get("vencimiento") or fecha_emi

    serie  = (factura.get("serie") or "").strip()
    numero = (factura.get("numero") or "").strip()
    if serie and numero:
        serie_numero = f"{serie}-{numero}"
    else:
        serie_numero = serie or numero

    concepto        = (factura.get("concepto") or "").strip()
    glosa_principal = concepto[:40]
    glosa_detalle   = concepto[:30]

    tasa = contab.get("tasa_igv", 18.0)

    def _fila_base(monto: float) -> Dict:
        is_usd = (moneda == "USD")
        is_pen = (moneda == "PEN")
        return {
            "A":  "",
            "B":  sub_diario,
            "C":  numero_comprobante,
            "D":  fecha_emi,
            "E":  moneda_code,
            "F":  glosa_principal,
            "G":  "",
            "H":  _DEFAULT_TIPO_CONVERSION,
            "I":  _DEFAULT_FLAG_CONVERSION,
            "J":  fecha_hoy,
            "K":  "",
            "L":  "",
            "M":  "",
            "N":  "",
            "O":  monto,
            "P":  monto if is_usd else "",
            "Q":  monto if is_pen else "",
            "R":  tipo_concar,
            "S":  serie_numero,
            "T":  fecha_emi,
            "U":  vencimiento,
            "V":  "",
            "W":  glosa_detalle,
            "X":  "",
            "Y":  "",
            "Z":  "",
            "AA": "",
            "AB": "",
            "AC": "",
            "AD": "",
            "AE": "",
            "AF": "",
            "AG": "",
            "AH": "",
            "AI": "",
            "AJ": "",
            "AK": "",
            "AL": "",
            "AM": "",
            "AN": "",
            "AO": tasa,
        }

    filas: List[Dict] = []

    # ── LÍNEA 1: GASTO (DEBE) ─────────────────────────────────────
    fila_gasto = _fila_base(base)
    fila_gasto["K"] = contab["cuentas"]["gasto"]
    fila_gasto["M"] = cc
    fila_gasto["N"] = "D"
    filas.append(fila_gasto)

    # ── LÍNEA 2: IGV CRÉDITO FISCAL (DEBE) — solo si hay IGV ──────
    if monto_igv > 0:
        fila_igv = _fila_base(monto_igv)
        fila_igv["K"] = contab["cuentas"]["igv"]
        fila_igv["F"] = f"IGV - {concepto}"[:40]
        fila_igv["W"] = f"IGV - {concepto}"[:30]
        fila_igv["N"] = "D"
        filas.append(fila_igv)

    # ── LÍNEA 3: CXP PROVEEDOR (HABER) ────────────────────────────
    fila_cxp = _fila_base(monto_total)
    fila_cxp["K"] = contab["cuentas"]["cxp"]
    fila_cxp["L"] = ruc
    fila_cxp["N"] = "H"
    filas.append(fila_cxp)

    return filas


# ─────────────────────────────────────────────────────────────────────────
# Generador del .xlsx final
# ─────────────────────────────────────────────────────────────────────────

def build_xlsx(filas: List[Dict]) -> bytes:
    """
    Construye un .xlsx en memoria con:
      - Fila 1: headers azules (cabecera de columna).
      - Fila 2: descripciones de validación.
      - Fila 3: tamaño/formato de cada columna.
      - Fila 4 en adelante: las filas generadas por factura_a_filas
        (keys = letras de columna).

    Devuelve el contenido binario del .xlsx listo para servir.
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CONCAR"

    # Estilos del template original
    fill_blue   = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    font_white  = Font(color="FFFFFF", bold=True)
    font_small  = Font(size=9, italic=True)
    align_wrap  = Alignment(wrap_text=True, vertical="top")

    # Escribir fila 1 (headers azules)
    for col, header in EXCEL_HEADERS["row1"].items():
        cell = ws[f"{col}1"]
        cell.value = header
        cell.fill = fill_blue
        cell.font = font_white
        cell.alignment = align_wrap

    # Escribir fila 2 (descripciones amarillas)
    for col, desc in EXCEL_HEADERS["row2"].items():
        cell = ws[f"{col}2"]
        cell.value = desc
        cell.fill = fill_yellow
        cell.font = font_small
        cell.alignment = align_wrap

    # Escribir fila 3 (formatos)
    for col, fmt in EXCEL_HEADERS["row3"].items():
        cell = ws[f"{col}3"]
        cell.value = fmt
        cell.font = font_small
        cell.alignment = align_wrap

    # Escribir datos desde fila 4
    for row_idx, fila in enumerate(filas, start=4):
        for col_letter, value in fila.items():
            ws[f"{col_letter}{row_idx}"] = value

    # Anchos de columna extraídos del template real
    anchos = {
        "A": 17.25,
        "B": 12.75,
        "C": 18.75,
        "D": 13.75,
        "E": 11.75,
        "F": 41.5,
        "G": 18.75,
        "H": 13.0,
        "I": 16.75,
        "J": 14.75,
        "K": 15.75,
        "L": 13.0,
        "M": 16.75,
        "N": 12.75,
        "O": 18.75,
        "P": 13.0,
        "Q": 13.0,
        "R": 17.75,
        "S": 13.0,
        "T": 16.75,
        "U": 13.0,
        "V": 15.75,
        "W": 41.13,
        "X": 15.75,
        "Y": 18.75,
        "Z": 14.75,
        "AA": 16.75,
        "AB": 14.75,
        "AC": 15.75,
        "AD": 13.0,
        "AE": 13.0,
        "AF": 19.75,
        "AG": 13.0,
        "AH": 13.0,
        "AI": 17.75,
        "AJ": 22.75,
        "AK": 13.0,
        "AL": 13.0,
        "AM": 18.75,
        "AN": 13.0,
        "AO": 13.0,
    }
    
    for col, width in anchos.items():
        ws.column_dimensions[col].width = width

    # Altura de filas de header (si se especificaron en el template)
    # ws.row_dimensions[1].height = 36
    # ws.row_dimensions[2].height = 60

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()